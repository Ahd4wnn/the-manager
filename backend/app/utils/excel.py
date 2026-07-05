"""Colourful, detailed .xlsx financial report via openpyxl."""

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.billing import Invoice
from app.models.expense import Expense
from app.models.hospital import Hospital
from app.services.reports import financial_summary, staff_performance

# Palette (matches the "Clinical Calm" theme)
ACCENT = "0071E3"
ACCENT_DARK = "0A3D91"
GREEN = "248A3D"
ORANGE = "B25000"
HEADER_TEXT = "FFFFFF"
BAND = "F2F6FC"
CARD = "EAF2FE"
INR_FMT = '"₹"#,##0.00'

_thin = Side(style="thin", color="D8DEE9")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header(cell, text, fill=ACCENT):
    cell.value = text
    cell.font = Font(bold=True, color=HEADER_TEXT, size=11)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def _title(ws, text, sub):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=18, color=ACCENT_DARK)
    ws["A2"] = sub
    ws["A2"].font = Font(size=11, color="6E6E73")


def build_financial_workbook(
    db: Session,
    hospital: Hospital,
    from_date: date,
    to_date: date,
    granularity: str = "daily",
) -> bytes:
    summary = financial_summary(db, hospital.id, from_date, to_date, granularity)
    perf = staff_performance(db, hospital.id, from_date, to_date)

    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    _title(ws, hospital.name, f"Financial report · {from_date} to {to_date}")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20

    kpis = [
        ("Revenue collected", summary.total_revenue, GREEN),
        ("Expenses", summary.total_expenses, ORANGE),
        ("Net", summary.net, ACCENT_DARK),
        ("Invoiced (period)", summary.invoiced, ACCENT),
        ("Outstanding (all)", summary.outstanding, ORANGE),
        ("New patients", Decimal(summary.patients), ACCENT),
        ("Invoices raised", Decimal(summary.invoices), ACCENT),
    ]
    row = 4
    for label, value, color in kpis:
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=True, color="1C1C1E")
        c.fill = PatternFill("solid", fgColor=CARD)
        c.border = BORDER
        v = ws.cell(row=row, column=2, value=float(value))
        v.font = Font(bold=True, size=12, color=color)
        v.border = BORDER
        if label not in ("New patients", "Invoices raised"):
            v.number_format = INR_FMT
        row += 1

    # ---- Trend sheet ----
    tw = wb.create_sheet("Trend")
    _title(tw, "Revenue vs Expenses", f"By {granularity}")
    headers = ["Period", "Revenue", "Expenses", "Net"]
    for i, h in enumerate(headers, start=1):
        _header(tw.cell(row=4, column=i), h)
        tw.column_dimensions[get_column_letter(i)].width = 18
    r = 5
    for b in summary.buckets:
        band = BAND if (r % 2 == 0) else "FFFFFF"
        vals = [b.label, float(b.revenue), float(b.expenses), float(b.net)]
        for i, val in enumerate(vals, start=1):
            cell = tw.cell(row=r, column=i, value=val)
            cell.fill = PatternFill("solid", fgColor=band)
            cell.border = BORDER
            if i > 1:
                cell.number_format = INR_FMT
        r += 1
    # totals row
    for i, val in enumerate(
        ["Total", float(summary.total_revenue), float(summary.total_expenses), float(summary.net)],
        start=1,
    ):
        cell = tw.cell(row=r, column=i, value=val)
        cell.font = Font(bold=True, color=HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=ACCENT_DARK)
        cell.border = BORDER
        if i > 1:
            cell.number_format = INR_FMT

    # ---- Expenses sheet ----
    ew = wb.create_sheet("Expenses")
    _title(ew, "Expenses", f"{from_date} to {to_date}")
    for i, h in enumerate(["Date", "Category", "Amount", "Note"], start=1):
        _header(ew.cell(row=4, column=i), h, fill=ORANGE)
    ew.column_dimensions["A"].width = 14
    ew.column_dimensions["B"].width = 20
    ew.column_dimensions["C"].width = 16
    ew.column_dimensions["D"].width = 40
    expenses = db.scalars(
        select(Expense).where(
            Expense.hospital_id == hospital.id,
            Expense.deleted_at.is_(None),
            Expense.spent_on >= from_date,
            Expense.spent_on <= to_date,
        ).order_by(Expense.spent_on)
    ).all()
    r = 5
    for e in expenses:
        band = BAND if (r % 2 == 0) else "FFFFFF"
        for i, val in enumerate(
            [e.spent_on.isoformat(), e.category, float(e.amount), e.note or ""], start=1
        ):
            cell = ew.cell(row=r, column=i, value=val)
            cell.fill = PatternFill("solid", fgColor=band)
            cell.border = BORDER
            if i == 3:
                cell.number_format = INR_FMT
        r += 1

    # ---- Invoices sheet ----
    iw = wb.create_sheet("Invoices")
    _title(iw, "Invoices", f"Raised {from_date} to {to_date}")
    for i, h in enumerate(
        ["Invoice", "Status", "Total", "Paid", "Balance"], start=1
    ):
        _header(iw.cell(row=4, column=i), h)
    for col, w in zip("ABCDE", (22, 16, 16, 16, 16)):
        iw.column_dimensions[col].width = w
    invoices = db.scalars(
        select(Invoice).where(
            Invoice.hospital_id == hospital.id, Invoice.deleted_at.is_(None)
        ).order_by(Invoice.created_at.desc())
    ).all()
    r = 5
    for inv in invoices:
        band = BAND if (r % 2 == 0) else "FFFFFF"
        vals = [
            inv.invoice_number,
            inv.status.value if hasattr(inv.status, "value") else str(inv.status),
            float(inv.total_amount),
            float(inv.amount_paid),
            float(inv.balance_due),
        ]
        for i, val in enumerate(vals, start=1):
            cell = iw.cell(row=r, column=i, value=val)
            cell.fill = PatternFill("solid", fgColor=band)
            cell.border = BORDER
            if i >= 3:
                cell.number_format = INR_FMT
        r += 1

    # ---- Staff performance sheet ----
    sw = wb.create_sheet("Staff")
    _title(sw, "Staff performance", f"{from_date} to {to_date}")
    for i, h in enumerate(
        ["Name", "Role", "Designation", "Treatments", "Collected"], start=1
    ):
        _header(sw.cell(row=4, column=i), h, fill=GREEN)
    for col, w in zip("ABCDE", (24, 12, 18, 14, 16)):
        sw.column_dimensions[col].width = w
    r = 5
    for s in perf.staff:
        band = BAND if (r % 2 == 0) else "FFFFFF"
        vals = [
            s.full_name,
            s.role,
            s.designation or "",
            s.treatments_performed,
            float(s.payments_collected),
        ]
        for i, val in enumerate(vals, start=1):
            cell = sw.cell(row=r, column=i, value=val)
            cell.fill = PatternFill("solid", fgColor=band)
            cell.border = BORDER
            if i == 5:
                cell.number_format = INR_FMT
        r += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
